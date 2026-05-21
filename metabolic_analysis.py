"""
Metabolic Analysis Pipeline — BW, TEI, Metabolic Profiling
Universal parser for longitudinal metabolic data. Reads Excel from Rawdata/.
"""
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy import stats
import openpyxl
import glob
import re
import os
from collections import OrderedDict

# ── Configuration ──────────────────────────────────────────────────────────
OUTPUT_DIR = r"C:\AI\projects\Metabolic\figure"
RAWDATA_DIR = r"C:\AI\projects\Metabolic\Rawdata"

STUDY_ID = None
raw_data_all = {}  # {sheet_name: [animal_dicts]}
groups_data_all = {}  # {sheet_name: {group: {metric: summary}}}
STAT_RESULTS_ALL = {}  # {sheet_name: {group: {key: {p: ...}}}}
GROUP_ORDER = []

GROUP_COLORS = {
    "Vehicle": "#377eb8", "Sema": "#ff7f00",
    "JKL-010": "#4daf4a", "JKL-010+Sema": "#984ea3",
}
DEFAULT_COLORS = ["#377eb8", "#ff7f00", "#4daf4a", "#984ea3",
                  "#a65628", "#f781bf", "#999999", "#e41a1c"]

plt.rcParams.update({
    "font.family": "sans-serif", "font.size": 9,
    "axes.titlesize": 10, "axes.labelsize": 9,
    "legend.fontsize": 7, "figure.dpi": 150, "savefig.dpi": 150,
    "savefig.bbox": "tight", "savefig.pad_inches": 0.15,
})

# ── Helpers ─────────────────────────────────────────────────────────────────
def pvalue_stars(p):
    if p < 0.001: return "***"
    elif p < 0.01: return "**"
    elif p < 0.05: return "*"
    return "ns"

def clean_group_name(raw):
    if not raw: return "Unknown"
    c = re.sub(r"^\d+\s*\n\s*", "", raw)
    c = re.sub(r"\s*\n\s*", "", c)
    return re.sub(r"\s+", " ", c).strip()

def find_data_file():
    candidates = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(glob.glob(os.path.join(RAWDATA_DIR, ext)))
    candidates = list(set(p for p in candidates if not os.path.basename(p).startswith("~$")))
    if not candidates:
        raise FileNotFoundError(f"No Excel file found in {RAWDATA_DIR}")
    return candidates[0]

def _save_fig(fig, name):
    path = os.path.join(OUTPUT_DIR, name)
    fig.savefig(path, dpi=150, bbox_inches="tight", pad_inches=0.15,
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"  Saved: {name}")

def _safe_float(v):
    """Convert cell value to float, return None if not possible."""
    if v is None: return None
    try: return float(v)
    except (ValueError, TypeError): return None

# ── Generic Sheet Parser ────────────────────────────────────────────────────
def parse_sheet(ws, sheet_name):
    """
    Generic parser for metabolic sheets.
    Detects header row, date/study-day rows, metric blocks, and data rows.
    Returns (raw_data, metric_blocks, time_labels, study_id).

    metric_blocks: OrderedDict of {block_name: {"start_col": int, "n_tp": int, "unit": str}}
    time_labels: list of study day labels (or None if not time-series)
    raw_data: list of dicts per animal
    """
    # Find header row (contains "Group" in col 1)
    header_row = None
    for row in range(1, ws.max_row + 1):
        v = str(ws.cell(row=row, column=1).value or "").strip()
        if v == "Group":
            header_row = row
            break
    if header_row is None:
        raise ValueError(f"Could not find 'Group' header in sheet {sheet_name}")

    # Find date/study-day row (look for "Date" or "Study Day" in col 3)
    date_row = None
    study_day_row = None
    for row in range(header_row + 1, header_row + 5):
        v = str(ws.cell(row=row, column=3).value or "").strip()
        if v == "Date":
            date_row = row
        elif v == "Study Day":
            study_day_row = row

    # Detect metric blocks from header row — handle repeated block names
    metric_blocks = OrderedDict()
    block_counts = {}
    current_block = None
    for col in range(4, ws.max_column + 1):
        hdr = str(ws.cell(row=header_row, column=col).value or "").strip()

        if hdr and hdr not in ("Group", "Animal ID", "Cage No.", "Cage"):
            # New block header found
            if hdr in block_counts:
                block_counts[hdr] += 1
                block_name = f"{hdr}_{block_counts[hdr]}"
            else:
                block_counts[hdr] = 1
                block_name = hdr
            current_block = block_name
            metric_blocks[current_block] = {"start_col": col, "n_tp": 0, "unit": hdr}
        elif current_block and not hdr:
            # Continuation of current block
            pass

    # Count timepoints per block: count non-empty cells in the sub-header or study-day row
    for name, info in metric_blocks.items():
        sc = info["start_col"]
        n = 0
        for col in range(sc, ws.max_column + 1):
            # Check if we've hit the next block
            next_hdr = str(ws.cell(row=header_row, column=col).value or "").strip()
            if next_hdr and col > sc:
                break
            # Count if sub-header or date/study-day has a value
            sub_val = ws.cell(row=header_row + 1, column=col).value
            sub_val2 = ws.cell(row=header_row + 2, column=col).value if header_row + 2 <= ws.max_row else None
            if sub_val is not None or sub_val2 is not None:
                n += 1
        info["n_tp"] = max(info["n_tp"], n)

    # Get time labels from study_day_row
    time_labels = []
    if study_day_row:
        for col in range(4, ws.max_column + 1):
            v = _safe_float(ws.cell(row=study_day_row, column=col).value)
            if v is not None and v not in time_labels:
                time_labels.append(int(v))
    if not time_labels:
        time_labels = list(range(max(b["n_tp"] for b in metric_blocks.values())))

    # Find data rows
    data_start = None
    first_data_row = study_day_row or date_row or (header_row + 1)
    for row in range(first_data_row + 1, ws.max_row + 1):
        id_val = str(ws.cell(row=row, column=2).value or "").strip()
        grp_val = str(ws.cell(row=row, column=1).value or "").strip()
        if id_val and id_val not in ("Animal ID", "Region", "Date", "Study Day"):
            data_start = row
            break

    if data_start is None:
        raise ValueError(f"Could not find data rows in {sheet_name}")

    data_end = data_start
    for row in range(data_start + 1, ws.max_row + 1):
        id_val = str(ws.cell(row=row, column=2).value or "").strip()
        if not id_val:
            break
        data_end = row

    # Parse animal data
    raw_data = []
    current_group_raw = None
    for row in range(data_start, data_end + 1):
        grp = ws.cell(row=row, column=1).value
        if grp and str(grp).strip():
            current_group_raw = str(grp).strip()
        group_display = clean_group_name(current_group_raw)

        animal_id = str(ws.cell(row=row, column=2).value or "").strip()
        if not animal_id:
            continue
        cage = str(ws.cell(row=row, column=3).value or "").strip()

        animal = {"group": group_display, "id": animal_id, "cage": cage}
        for name, info in metric_blocks.items():
            values = []
            for offset in range(info["n_tp"]):
                v = _safe_float(ws.cell(row=row, column=info["start_col"] + offset).value)
                values.append(v)
            animal[name] = values
        raw_data.append(animal)

    return raw_data, metric_blocks, time_labels


# ── Compute derived metrics ─────────────────────────────────────────────────
def compute_pct_change(raw_data, metric_blocks):
    """Compute % change from first timepoint for each metric."""
    for d in raw_data:
        for name, info in metric_blocks.items():
            vals = d[name]
            base = vals[0] if vals[0] and vals[0] != 0 else None
            if base:
                d[f"{name}_pct"] = [(v - base) / base * 100 if v is not None else None
                                     for v in vals]
            else:
                d[f"{name}_pct"] = [None] * len(vals)


# ── Plotting functions ──────────────────────────────────────────────────────
def plot_time_courses(raw_data, metric_blocks, time_labels, sheet_name, group_order):
    """Generate time-course line plots for each metric block (skip single-timepoint)."""
    tp_blocks = OrderedDict((k, v) for k, v in metric_blocks.items() if v["n_tp"] > 1)
    if not tp_blocks:
        return
    n_metrics = len(tp_blocks)
    n_cols = min(4, n_metrics)
    n_rows = int(np.ceil(n_metrics / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4.5 * n_rows),
                             squeeze=False)

    for idx, (name, info) in enumerate(tp_blocks.items()):
        ax = axes[idx // n_cols][idx % n_cols]
        n_tp = info["n_tp"]
        tl = time_labels[:n_tp] if len(time_labels) >= n_tp else list(range(n_tp))

        for g in group_order:
            members = [d for d in raw_data if d["group"] == g]
            all_vals = []
            for d in members:
                vals = [v for v in d[name][:n_tp] if v is not None]
                if len(vals) == n_tp:
                    all_vals.append(vals)
            if not all_vals:
                continue
            all_vals = np.array(all_vals)
            mean = np.mean(all_vals, axis=0)
            sem = np.std(all_vals, axis=0, ddof=1) / np.sqrt(len(all_vals))
            ax.fill_between(tl[:len(mean)], mean - sem, mean + sem,
                            color=GROUP_COLORS.get(g, "#999"), alpha=0.15)
            ax.plot(tl[:len(mean)], mean, "-", color=GROUP_COLORS.get(g, "#999"),
                    linewidth=2, label=g)

        ax.set_title(name, fontsize=9, fontweight="bold")
        ax.set_xlabel("Study Day")
        ax.legend(fontsize=7, loc="upper left")

    # Hide unused subplots
    for idx in range(n_metrics, n_rows * n_cols):
        axes[idx // n_cols][idx % n_cols].set_visible(False)

    fig.suptitle(f"{STUDY_ID}: {sheet_name} — Time Courses (Mean ± SEM)",
                 fontsize=12, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    _save_fig(fig, f"{sheet_name.lower().replace(' ', '_')}_time_courses.png")


def plot_pct_change_time_courses(raw_data, metric_blocks, time_labels, sheet_name, group_order):
    """% Change from baseline time-course plots (skip single-timepoint)."""
    tp_blocks = OrderedDict((k, v) for k, v in metric_blocks.items() if v["n_tp"] > 1)
    if not tp_blocks:
        return
    n_metrics = len(tp_blocks)
    n_cols = min(4, n_metrics)
    n_rows = int(np.ceil(n_metrics / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4.5 * n_rows),
                             squeeze=False)

    for idx, (name, info) in enumerate(tp_blocks.items()):
        ax = axes[idx // n_cols][idx % n_cols]
        n_tp = info["n_tp"]
        tl = time_labels[:n_tp] if len(time_labels) >= n_tp else list(range(n_tp))

        for g in group_order:
            members = [d for d in raw_data if d["group"] == g]
            all_pct = []
            for d in members:
                pct_key = f"{name}_pct"
                vals = [v for v in d.get(pct_key, [])[:n_tp] if v is not None]
                if len(vals) == n_tp:
                    all_pct.append(vals)
            if not all_pct:
                continue
            all_pct = np.array(all_pct)
            mean = np.mean(all_pct, axis=0)
            sem = np.std(all_pct, axis=0, ddof=1) / np.sqrt(len(all_pct))
            ax.fill_between(tl[:len(mean)], mean - sem, mean + sem,
                            color=GROUP_COLORS.get(g, "#999"), alpha=0.15)
            ax.plot(tl[:len(mean)], mean, "-", color=GROUP_COLORS.get(g, "#999"),
                    linewidth=2, label=g)

        ax.axhline(y=0, color="black", linewidth=0.5, linestyle="--")
        ax.set_title(f"{name} — % Change", fontsize=9, fontweight="bold")
        ax.set_xlabel("Study Day")
        ax.set_ylabel("% Change from Day 0")
        ax.legend(fontsize=7, loc="upper left")

    for idx in range(n_metrics, n_rows * n_cols):
        axes[idx // n_cols][idx % n_cols].set_visible(False)

    fig.suptitle(f"{STUDY_ID}: {sheet_name} — % Change from Baseline (Mean ± SEM)",
                 fontsize=12, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    _save_fig(fig, f"{sheet_name.lower().replace(' ', '_')}_pct_change_curves.png")


def plot_final_pct_bars(raw_data, metric_blocks, sheet_name, group_order):
    """
    Bar charts of % change at the final timepoint for each metric.
    Uses vs-0 t-test and between-group (vs Vehicle) significance.
    """
    pct_metrics = OrderedDict()
    for name, info in metric_blocks.items():
        if info["n_tp"] >= 2:
            pct_metrics[name] = info

    if not pct_metrics:
        return

    n_metrics = len(pct_metrics)
    n_cols = min(4, n_metrics)
    n_rows = int(np.ceil(n_metrics / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 5 * n_rows),
                             squeeze=False)
    x = np.arange(len(group_order))
    bar_width = 0.45

    for idx, (name, info) in enumerate(pct_metrics.items()):
        ax = axes[idx // n_cols][idx % n_cols]
        last_idx = info["n_tp"] - 1

        # Compute % change at final timepoint for each animal
        group_vals = {}
        for g in group_order:
            pct_vals = []
            for d in raw_data:
                if d["group"] != g:
                    continue
                base = d[name][0]
                final = d[name][last_idx]
                if base and base != 0 and final is not None:
                    pct_vals.append((final - base) / base * 100)
            group_vals[g] = pct_vals

        means = np.array([np.mean(group_vals[g]) if group_vals[g] else 0
                          for g in group_order])
        sems = np.array([np.std(group_vals[g], ddof=1) / np.sqrt(len(group_vals[g]))
                         if len(group_vals[g]) > 1 else 0 for g in group_order])
        colors = [GROUP_COLORS.get(g, "#999") for g in group_order]

        ax.bar(x, means, bar_width, yerr=sems, capsize=6,
               color=colors, alpha=0.85, edgecolor="white", linewidth=0.8)
        ax.axhline(y=0, color="black", linewidth=0.8)

        y_span = max(abs(np.min(means - sems)), abs(np.max(means + sems))) * 2 + 0.01
        y_buf = y_span * 0.06
        tracked_max = max(0, np.max(means + sems))
        tracked_min = min(0, np.min(means - sems))

        # Vs-0 stars
        for i, g in enumerate(group_order):
            vals = group_vals[g]
            if len(vals) < 2:
                continue
            t_stat, p_val = stats.ttest_1samp(vals, 0)
            stars = pvalue_stars(p_val)
            if stars == "ns":
                continue
            sign = 1 if means[i] > 0 else -1
            y_pos = means[i] + sign * (sems[i] + y_buf * 1.5)
            ax.text(x[i], y_pos, stars, ha="center",
                    va="bottom" if sign > 0 else "top",
                    fontsize=8, fontweight="bold")
            if y_pos > 0:
                tracked_max = max(tracked_max, y_pos + y_buf)
            else:
                tracked_min = min(tracked_min, y_pos - y_buf)

        # Between-group brackets (vs Vehicle)
        veh_vals = group_vals.get("Vehicle", [])
        bracket_level = 0
        for i, g in enumerate(group_order):
            if g == "Vehicle" or not veh_vals or len(group_vals[g]) < 2:
                continue
            t_stat, p_val = stats.ttest_ind(group_vals[g], veh_vals, equal_var=False)
            stars = pvalue_stars(p_val)
            if stars == "ns":
                continue
            top_y = max(means[0] + sems[0], means[i] + sems[i])
            h = y_buf * (4 + bracket_level * 1.8)
            bracket_y = top_y + h
            ax.plot([x[0], x[0], x[i], x[i]],
                    [bracket_y, bracket_y + y_buf * 0.3,
                     bracket_y + y_buf * 0.3, bracket_y],
                    lw=1.0, color="gray", clip_on=False)
            ax.text((x[0] + x[i]) / 2, bracket_y + y_buf * 0.35,
                    stars, ha="center", va="bottom",
                    fontsize=7, fontweight="bold", color="gray")
            tracked_max = max(tracked_max, bracket_y + y_buf * 0.8)
            bracket_level += 1

        margin = y_span * 0.12
        ax.set_ylim(tracked_min - margin, tracked_max + margin)
        ax.set_xticks(x)
        ax.set_xticklabels(group_order, fontsize=8)
        ax.set_title(f"{name}", fontsize=9, fontweight="bold")
        if idx % n_cols == 0:
            ax.set_ylabel("% Change from Baseline")

    for idx in range(n_metrics, n_rows * n_cols):
        axes[idx // n_cols][idx % n_cols].set_visible(False)

    fig.suptitle(f"{STUDY_ID}: {sheet_name} — % Change at Final Timepoint\n"
                 "* p<0.05, ** p<0.01, *** p<0.001 vs 0 (black) | vs Vehicle (gray)",
                 fontsize=11, fontweight="bold")
    fig.tight_layout(rect=[0, 0.02, 1, 0.92])
    _save_fig(fig, f"{sheet_name.lower().replace(' ', '_')}_pct_bars.png")


# ── Process a single sheet ──────────────────────────────────────────────────
def process_sheet(ws, sheet_name, group_order, study_id):
    """Parse, analyze, and plot for one sheet."""
    print(f"\n{'='*60}")
    print(f"Processing: {sheet_name}")
    print(f"{'='*60}")

    raw_data, metric_blocks, time_labels = parse_sheet(ws, sheet_name)
    compute_pct_change(raw_data, metric_blocks)

    n_animals = len(raw_data)
    print(f"  Animals: {n_animals}")
    print(f"  Groups: {group_order}")
    safe_names = [str(k).encode('ascii', errors='replace').decode('ascii')
                  for k in metric_blocks.keys()]
    print(f"  Metrics: {safe_names}")

    for name, info in metric_blocks.items():
        safe_name = str(name).encode('ascii', errors='replace').decode('ascii')
        print(f"    {safe_name}: {info['n_tp']} timepoints")

    # Time-course plots
    if any(info["n_tp"] > 1 for info in metric_blocks.values()):
        plot_time_courses(raw_data, metric_blocks, time_labels, sheet_name, group_order)
        plot_pct_change_time_courses(raw_data, metric_blocks, time_labels, sheet_name, group_order)

    # Final % change bar charts
    plot_final_pct_bars(raw_data, metric_blocks, sheet_name, group_order)

    return raw_data, metric_blocks


# ── Main ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    filepath = find_data_file()
    STUDY_ID = os.path.splitext(os.path.basename(filepath))[0].split("-")[0]

    print(f"Metabolic Analysis — {STUDY_ID}")
    print(f"Reading: {os.path.basename(filepath)}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    target_sheets = ["BW", "TEI", "Metabolic Profiling"]
    available = [s for s in target_sheets if s in wb.sheetnames]
    print(f"Available sheets: {available}")

    # First, parse one sheet to get group order
    first_sheet = available[0]
    raw_data, metric_blocks, _ = parse_sheet(wb[first_sheet], first_sheet)
    GROUP_ORDER = []
    for d in raw_data:
        if d["group"] not in GROUP_ORDER:
            GROUP_ORDER.append(d["group"])

    for i, g in enumerate(GROUP_ORDER):
        if g not in GROUP_COLORS:
            GROUP_COLORS[g] = DEFAULT_COLORS[i % len(DEFAULT_COLORS)]

    print(f"\nGroup order: {GROUP_ORDER}")
    print(f"Output: {OUTPUT_DIR}")

    # Process each sheet
    for sheet_name in available:
        process_sheet(wb[sheet_name], sheet_name, GROUP_ORDER, STUDY_ID)

    wb.close()

    print(f"\n{'='*60}")
    print(f"Done! All outputs in: {OUTPUT_DIR}")
